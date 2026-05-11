"""Create or edit an .xlsx from the office worker's sheets/ops DSL (openpyxl).

Usage:
  python3 xlsx_build.py create <output.xlsx> <json-spec>
  python3 xlsx_build.py edit   <input.xlsx> <output.xlsx> <json-ops>

`json-spec` = {"sheets": Sheet[]} with optional native `charts`; `json-ops` = Op[].
Output: {"ok": true, "path": "<output>"}
"""

from __future__ import annotations

import json

from script_runner import main, ok, fail


def _chart_class(kind: str):
    """Map a DSL chart kind to an openpyxl chart class."""
    from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart

    return {"bar": BarChart, "line": LineChart, "pie": PieChart, "scatter": ScatterChart}.get(kind)


def _add_chart(ws, chart_spec: dict) -> None:
    """Add a native chart to worksheet ``ws`` from a DSL chart spec, bound to cell ranges."""
    from openpyxl.chart import Reference

    cls = _chart_class(chart_spec.get("type", ""))
    if cls is None:
        fail(f"unknown chart type: {chart_spec.get('type')}")
    chart = cls()
    if chart_spec.get("title"):
        chart.title = str(chart_spec["title"])

    def parse_ref(s: str) -> Reference:
        # Accept "Sheet!A1:B10" or "A1:B10" (defaults to ws).
        if "!" in s:
            sheet_name, rng = s.split("!", 1)
            target = ws.parent[sheet_name]
        else:
            target, rng = ws, s
        a1 = rng.split(":")
        if len(a1) != 2:
            fail(f"invalid range: {s}")
        from openpyxl.utils.cell import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(rng)
        return Reference(target, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)

    data = parse_ref(str(chart_spec["dataRange"]))
    chart.add_data(data, titles_from_data=False)
    if chart_spec.get("categoriesRange"):
        chart.set_categories(parse_ref(str(chart_spec["categoriesRange"])))
    ws.add_chart(chart, str(chart_spec["anchor"]))


def _write_sheet(wb, sheet_spec: dict, *, first: bool) -> None:
    """Create (or, for the first one, reuse the default) a worksheet from a DSL sheet spec."""
    name = str(sheet_spec["name"])
    ws = wb.active if first else wb.create_sheet()
    ws.title = name
    for row in sheet_spec.get("rows", []):
        ws.append(list(row))
    if sheet_spec.get("freeze"):
        ws.freeze_panes = str(sheet_spec["freeze"])
    for chart_spec in sheet_spec.get("charts", []) or []:
        _add_chart(ws, chart_spec)


def _create(output: str, spec: dict) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    sheets = spec.get("sheets", [])
    if not sheets:
        fail("spec.sheets must be non-empty")
    for i, sheet_spec in enumerate(sheets):
        _write_sheet(wb, sheet_spec, first=(i == 0))
    wb.save(output)
    ok(path=output)


def _edit(src: str, output: str, ops: list) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(src)
    for op in ops:
        kind = op.get("op")
        if kind == "set_cell":
            ws = wb[str(op["sheet"])]
            ws[str(op["cell"])] = op["value"]
        elif kind == "set_formula":
            ws = wb[str(op["sheet"])]
            formula = str(op["formula"])
            ws[str(op["cell"])] = formula if formula.startswith("=") else f"={formula}"
        elif kind == "add_sheet":
            wb.create_sheet(title=str(op["name"]))
        elif kind == "add_chart":
            ws = wb[str(op["sheet"])]
            _add_chart(ws, op["chart"])
        else:
            fail(f"unknown op: {kind}")
    wb.save(output)
    ok(path=output)


def _run(argv: list[str]) -> None:
    if not argv:
        fail("usage: xlsx_build.py create|edit ...")
    mode = argv[0]
    if mode == "create":
        if len(argv) != 3:
            fail("usage: xlsx_build.py create <output.xlsx> <json-spec>")
        _create(argv[1], json.loads(argv[2]))
    elif mode == "edit":
        if len(argv) != 4:
            fail("usage: xlsx_build.py edit <input.xlsx> <output.xlsx> <json-ops>")
        _edit(argv[1], argv[2], json.loads(argv[3]))
    else:
        fail(f"unknown mode: {mode}")


if __name__ == "__main__":
    main(_run)
